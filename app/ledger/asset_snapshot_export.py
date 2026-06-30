from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TOTAL_FILL = PatternFill("solid", fgColor="E2F0D9")
EXCHANGE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B8C2CC"),
)
MONEY_FORMAT = "#,##0.00"


def _number(value):
    return float(value or 0)


def build_asset_snapshot_workbook(snapshots, matrix_builder):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "资产快照"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"

    current_row = 1
    max_columns = 7
    for snapshot in snapshots:
        (
            members,
            rows,
            currency_totals,
            base_total_row,
            exchange_gain_row,
            _grand_total,
        ) = matrix_builder(snapshot)
        rows = list(rows)
        total_columns = 3 + len(members) * 2 + 2
        max_columns = max(max_columns, total_columns)
        status_label = "（草稿）" if snapshot.is_draft else ""

        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=3,
        )
        worksheet.cell(
            current_row,
            1,
            f"{snapshot.snapshot_date:%Y年%m月%d日} 资产快照{status_label}",
        )
        for index, member in enumerate(members):
            start_column = 4 + index * 2
            worksheet.merge_cells(
                start_row=current_row,
                start_column=start_column,
                end_row=current_row,
                end_column=start_column + 1,
            )
            worksheet.cell(current_row, start_column, member.display_name)
        total_start = 4 + len(members) * 2
        worksheet.merge_cells(
            start_row=current_row,
            start_column=total_start,
            end_row=current_row,
            end_column=total_start + 1,
        )
        worksheet.cell(current_row, total_start, "合计")
        for cell in worksheet[current_row][:total_columns]:
            cell.fill = TITLE_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[current_row].height = 24

        header_row = current_row + 1
        headers = ["账户名称", "资产类别", "币种"]
        for _member in members:
            headers.extend(["原币", "本位币"])
        headers.extend(["原币", "本位币"])
        for column, value in enumerate(headers, start=1):
            cell = worksheet.cell(header_row, column, value)
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True, color="17365D")
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        write_row = header_row + 1
        for row in rows:
            values = [row["account_name"], row["asset_category"], row["currency"]]
            for cell in row["cells"]:
                if cell:
                    values.extend(
                        [_number(cell.original_amount), _number(cell.base_amount)]
                    )
                else:
                    values.extend([None, None])
            values.extend(
                [_number(row["total_original"]), _number(row["total_base"])]
            )
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(write_row, column, value)
                cell.border = THIN_BORDER
                if column >= 4:
                    cell.number_format = MONEY_FORMAT
                    cell.alignment = Alignment(horizontal="right")
            write_row += 1

        for total in currency_totals:
            values = [total["label"], None, None]
            for cell in total["cells"]:
                values.extend([_number(cell["original"]), _number(cell["base"])])
            values.extend(
                [_number(total["total_original"]), _number(total["total_base"])]
            )
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(write_row, column, value)
                cell.fill = TOTAL_FILL
                cell.border = THIN_BORDER
                if column >= 4:
                    cell.number_format = MONEY_FORMAT
            worksheet.merge_cells(
                start_row=write_row,
                start_column=1,
                end_row=write_row,
                end_column=3,
            )
            write_row += 1

        base_values = [base_total_row["label"], None, None]
        for amount in base_total_row["cells"]:
            base_values.extend(["人民币", _number(amount)])
        base_values.extend(["人民币", _number(base_total_row["total"])])
        for column, value in enumerate(base_values, start=1):
            cell = worksheet.cell(write_row, column, value)
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            if column >= 4 and isinstance(value, (int, float)):
                cell.number_format = MONEY_FORMAT
        worksheet.merge_cells(
            start_row=write_row,
            start_column=1,
            end_row=write_row,
            end_column=3,
        )
        write_row += 1

        exchange_values = [exchange_gain_row["label"], None, None]
        for amount in exchange_gain_row["cells"]:
            exchange_values.extend(["人民币", _number(amount)])
        exchange_values.extend(["人民币", _number(exchange_gain_row["total"])])
        for column, value in enumerate(exchange_values, start=1):
            cell = worksheet.cell(write_row, column, value)
            cell.fill = EXCHANGE_FILL
            cell.font = Font(bold=True, color="7F6000")
            cell.border = THIN_BORDER
            if column >= 4 and isinstance(value, (int, float)):
                cell.number_format = MONEY_FORMAT
        worksheet.merge_cells(
            start_row=write_row,
            start_column=1,
            end_row=write_row,
            end_column=3,
        )
        current_row = write_row + 3

    widths = {1: 24, 2: 18, 3: 10}
    for column in range(4, max_columns + 1):
        widths[column] = 15
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

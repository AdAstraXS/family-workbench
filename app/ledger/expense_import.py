from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
import json
from pathlib import Path

from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from family_core.models import FamilyMember
from portfolio.models import VisibilityChoices

from .models import BankAccount, ExpenseCategory, ExpenseImportBatch, ExpenseRecord


EXPECTED_HEADERS = (
    "支出时间",
    "所属账户",
    "支出账户",
    "一级分类",
    "二级分类",
    "三级分类",
    "金额",
    "备注",
)
ALLOWED_ACCOUNT_TYPE_CODES = ("bank", "wechat", "alipay")
MAX_IMPORT_ROWS = 20_000


class ExpenseWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class ParsedExpenseRow:
    row_number: int
    occurred_at: datetime
    member_name: str
    account_name: str
    category_names: tuple[str, ...]
    amount: Decimal
    remark: str


@dataclass(frozen=True)
class ExpenseImportResult:
    batch: ExpenseImportBatch
    duplicate_file: bool = False


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_occurred_at(value, workbook_epoch, row_number: int) -> datetime:
    parsed = None
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time.min)
    elif isinstance(value, (int, float, Decimal)):
        parsed = from_excel(value, workbook_epoch)
    elif value:
        text_value = _text(value)
        parsed = parse_datetime(text_value)
        if parsed is None:
            parsed_date = parse_date(text_value)
            if parsed_date:
                parsed = datetime.combine(parsed_date, time.min)
        if parsed is None:
            for date_format in (
                "%Y/%m/%d %H:%M:%S",
                "%Y/%m/%d %H:%M",
                "%Y/%m/%d",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M",
            ):
                try:
                    parsed = datetime.strptime(text_value, date_format)
                    break
                except ValueError:
                    continue
    if parsed is None:
        raise ExpenseWorkbookError(f"第 {row_number} 行：支出时间无法识别。")
    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed.astimezone(timezone.get_current_timezone()).replace(microsecond=0)


def _parse_amount(value, row_number: int) -> Decimal:
    try:
        amount = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        raise ExpenseWorkbookError(f"第 {row_number} 行：金额无法识别。") from None
    if not amount.is_finite():
        raise ExpenseWorkbookError(f"第 {row_number} 行：金额必须是有限数字。")
    if amount.as_tuple().exponent < -4:
        raise ExpenseWorkbookError(f"第 {row_number} 行：金额最多保留 4 位小数。")
    return amount


def parse_expense_workbook(file_bytes: bytes) -> tuple[str, list[ParsedExpenseRow]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ExpenseWorkbookError(f"无法读取 Excel 文件：{exc}") from exc

    if len(workbook.sheetnames) != 1:
        raise ExpenseWorkbookError("固定格式支出文件必须且只能包含 1 个工作表。")
    worksheet = workbook[workbook.sheetnames[0]]
    if worksheet.max_row > MAX_IMPORT_ROWS + 1:
        workbook.close()
        raise ExpenseWorkbookError(f"Excel 文件不能超过 {MAX_IMPORT_ROWS} 条支出记录。")
    actual_headers = tuple(_text(cell.value) for cell in worksheet[1][: len(EXPECTED_HEADERS)])
    if actual_headers != EXPECTED_HEADERS:
        expected = "、".join(EXPECTED_HEADERS)
        raise ExpenseWorkbookError(f"首行表头不符合固定格式，应依次为：{expected}。")

    worksheet_name = worksheet.title
    rows = []
    for row_number, cells in enumerate(
        worksheet.iter_rows(min_row=2, max_col=len(EXPECTED_HEADERS), values_only=True),
        start=2,
    ):
        if all(value in (None, "") for value in cells):
            continue
        occurred_at_value, member_value, account_value, primary_value, secondary_value, tertiary_value, amount_value, remark_value = cells
        member_name = _text(member_value)
        account_name = _text(account_value)
        primary_name = _text(primary_value)
        secondary_name = _text(secondary_value)
        tertiary_name = _text(tertiary_value)
        missing = [
            label
            for label, value in (
                ("所属账户", member_name),
                ("支出账户", account_name),
                ("一级分类", primary_name),
                ("二级分类", secondary_name),
            )
            if not value
        ]
        if missing:
            raise ExpenseWorkbookError(f"第 {row_number} 行：{ '、'.join(missing) }不能为空。")
        category_names = tuple(
            category_name
            for category_name in (primary_name, secondary_name, tertiary_name)
            if category_name
        )
        rows.append(
            ParsedExpenseRow(
                row_number=row_number,
                occurred_at=_parse_occurred_at(occurred_at_value, workbook.epoch, row_number),
                member_name=member_name,
                account_name=account_name,
                category_names=category_names,
                amount=_parse_amount(amount_value, row_number),
                remark=_text(remark_value),
            )
        )
    workbook.close()
    if not rows:
        raise ExpenseWorkbookError("Excel 文件中没有可导入的支出记录。")
    return worksheet_name, rows


def _fingerprint(row: ParsedExpenseRow) -> str:
    payload = {
        "occurred_at": row.occurred_at.isoformat(),
        "member": row.member_name,
        "account": row.account_name,
        "categories": row.category_names,
        "amount": format(row.amount.normalize(), "f"),
        "remark": row.remark,
    }
    return sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()


def _category_for_path(family, category_names, cache):
    parent = None
    for category_name in category_names:
        cache_key = (parent.pk if parent else None, category_name)
        category = cache.get(cache_key)
        if category is None:
            category, _ = ExpenseCategory.objects.get_or_create(
                family=family,
                parent=parent,
                name=category_name,
                defaults={"is_active": True},
            )
            if not category.is_active:
                category.is_active = True
                category.save(update_fields=["is_active"])
            cache[cache_key] = category
        parent = category
    return parent


def import_expense_workbook(*, family, uploaded_file, imported_by=None) -> ExpenseImportResult:
    original_position = uploaded_file.tell() if hasattr(uploaded_file, "tell") else None
    file_bytes = uploaded_file.read()
    if original_position is not None and hasattr(uploaded_file, "seek"):
        uploaded_file.seek(original_position)
    if not file_bytes:
        raise ExpenseWorkbookError("上传的 Excel 文件为空。")

    source_filename = Path(getattr(uploaded_file, "name", "支出导入.xlsx")).name
    source_sha256 = sha256(file_bytes).hexdigest()
    existing_batch = ExpenseImportBatch.objects.filter(
        family=family,
        source_sha256=source_sha256,
    ).first()
    if existing_batch:
        return ExpenseImportResult(batch=existing_batch, duplicate_file=True)

    worksheet_name, parsed_rows = parse_expense_workbook(file_bytes)
    member_map = {
        member.display_name: member
        for member in FamilyMember.objects.filter(family=family, is_active=True)
    }
    account_map = {
        (account.member.display_name, account.account_name): account
        for account in BankAccount.objects.filter(
            family=family,
            is_active=True,
            account_type_ref__code__in=ALLOWED_ACCOUNT_TYPE_CODES,
            account_type_ref__is_active=True,
        ).select_related("member", "account_type_ref")
    }

    for row in parsed_rows:
        if row.member_name not in member_map:
            raise ExpenseWorkbookError(f"第 {row.row_number} 行：找不到家庭成员“{row.member_name}”。")
        if (row.member_name, row.account_name) not in account_map:
            raise ExpenseWorkbookError(
                f"第 {row.row_number} 行：找不到“{row.member_name}”名下可用的"
                f"银行、微信或支付宝账户“{row.account_name}”。"
            )

    fingerprints = [_fingerprint(row) for row in parsed_rows]
    existing_fingerprints = set(
        ExpenseRecord.objects.filter(
            family=family,
            import_fingerprint__in=fingerprints,
        ).values_list("import_fingerprint", flat=True)
    )

    with transaction.atomic():
        actor = imported_by if getattr(imported_by, "is_authenticated", False) else None
        batch = ExpenseImportBatch.objects.create(
            family=family,
            imported_by=actor,
            source_filename=source_filename,
            source_sha256=source_sha256,
            worksheet_name=worksheet_name,
            row_count=len(parsed_rows),
            imported_count=0,
            skipped_count=0,
            total_amount=sum((row.amount for row in parsed_rows), Decimal("0")),
            extra_data={"expected_headers": list(EXPECTED_HEADERS)},
        )
        category_cache = {}
        records = []
        skipped_count = 0
        seen_fingerprints = set(existing_fingerprints)
        for row, fingerprint in zip(parsed_rows, fingerprints):
            if fingerprint in seen_fingerprints:
                skipped_count += 1
                continue
            seen_fingerprints.add(fingerprint)
            member = member_map[row.member_name]
            account = account_map[(row.member_name, row.account_name)]
            category = _category_for_path(family, row.category_names, category_cache)
            local_date = timezone.localtime(row.occurred_at).date()
            records.append(
                ExpenseRecord(
                    created_by=actor,
                    updated_by=actor,
                    family=family,
                    member=member,
                    bank_account=account,
                    category=category,
                    expense_date=local_date,
                    occurred_at=row.occurred_at,
                    period_start=local_date,
                    period_end=local_date,
                    amount=row.amount,
                    currency="CNY",
                    merchant="",
                    payment_method=account.account_type_ref.name,
                    visibility=VisibilityChoices.PRIVATE,
                    remark=row.remark,
                    import_batch=batch,
                    import_row_number=row.row_number,
                    import_fingerprint=fingerprint,
                    extra_data={
                        "source_categories": list(row.category_names),
                        "source_worksheet": worksheet_name,
                    },
                )
            )
        ExpenseRecord.objects.bulk_create(records, batch_size=500)
        batch.imported_count = len(records)
        batch.skipped_count = skipped_count
        batch.save(update_fields=["imported_count", "skipped_count", "updated_at"])
    return ExpenseImportResult(batch=batch)

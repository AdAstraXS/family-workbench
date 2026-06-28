from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from django.db import transaction
from django.db.models import Sum
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from family_core.models import AccountRegion, AccountType, AssetCategory, FamilyMember

from .models import AssetBalanceEntry, AssetBalanceSnapshot, BankAccount


WORKSHEET_NAME = "账户余额NEW"
MONEY_QUANTUM = Decimal("0.0001")
RATE_QUANTUM = Decimal("0.00000001")
CURRENCY_ALIASES = {"RMB": "CNY", "CNY": "CNY", "USD": "USD", "HKD": "HKD"}
ACCOUNT_ALIASES = {
    "粤商719056": "粤商证券1 xiao 719056",
    "粤商719800": "粤商证券2 xiao 719800",
    "粤商719775": "粤商证券 feng 719775",
    "粤商719777": "粤商证券 long 719777",
}


class AssetSnapshotWorkbookError(ValueError):
    pass


def _decimal(value, quantum=MONEY_QUANTUM):
    if value in (None, ""):
        value = 0
    return Decimal(str(value)).quantize(quantum, rounding=ROUND_HALF_UP)


@dataclass(frozen=True)
class ParsedAssetEntry:
    source_row: int
    member_name: str
    account_name: str
    region_name: str
    account_type_name: str
    asset_category_name: str
    currency: str
    original_amount: Decimal
    base_amount: Decimal


@dataclass(frozen=True)
class ParsedAssetSnapshot:
    snapshot_date: date
    usd_to_base: Decimal
    hkd_to_base: Decimal
    entries: tuple[ParsedAssetEntry, ...]

    @property
    def total_base(self):
        return sum((entry.base_amount for entry in self.entries), Decimal("0")).quantize(MONEY_QUANTUM)

    @property
    def member_totals(self):
        totals = {}
        for entry in self.entries:
            totals.setdefault(entry.member_name, Decimal("0"))
            totals[entry.member_name] += entry.base_amount
        return {name: amount.quantize(MONEY_QUANTUM) for name, amount in totals.items()}


@dataclass
class AssetSnapshotImportResult:
    created_dates: list[date] = field(default_factory=list)
    verified_dates: list[date] = field(default_factory=list)
    historical_accounts_created: list[str] = field(default_factory=list)
    snapshot_summaries: list[dict] = field(default_factory=list)


def parse_asset_snapshot_workbook(source):
    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise AssetSnapshotWorkbookError(f"找不到工作表“{WORKSHEET_NAME}”")
        worksheet = workbook[WORKSHEET_NAME]
        expected_headers = ("账户名称", "账户地区", "账户类型", "资产类别", "币种")
        actual_headers = tuple(worksheet.cell(4, column).value for column in range(1, 6))
        if actual_headers != expected_headers:
            raise AssetSnapshotWorkbookError(
                f"“{WORKSHEET_NAME}”第 4 行表头不符合固定格式：{actual_headers}"
            )

        snapshots = []
        for start_column in range(6, worksheet.max_column + 1, 4):
            raw_date = worksheet.cell(1, start_column).value
            if raw_date in (None, ""):
                continue
            if isinstance(raw_date, datetime):
                snapshot_date = raw_date.date()
            elif isinstance(raw_date, date):
                snapshot_date = raw_date
            elif isinstance(raw_date, (int, float, Decimal)):
                snapshot_date = from_excel(raw_date, workbook.epoch).date()
            else:
                raise AssetSnapshotWorkbookError(
                    f"第 1 行第 {start_column} 列不是有效快照日期：{raw_date}"
                )

            usd_to_base = _decimal(worksheet.cell(2, start_column + 1).value, RATE_QUANTUM)
            hkd_to_base = _decimal(worksheet.cell(2, start_column + 3).value, RATE_QUANTUM)
            member_columns = (
                (start_column, str(worksheet.cell(3, start_column).value or "").strip()),
                (start_column + 2, str(worksheet.cell(3, start_column + 2).value or "").strip()),
            )
            entries = []
            for row_number in range(5, worksheet.max_row + 1):
                account_name = str(worksheet.cell(row_number, 1).value or "").strip()
                region_name = str(worksheet.cell(row_number, 2).value or "").strip()
                account_type_name = str(worksheet.cell(row_number, 3).value or "").strip()
                asset_category_name = str(worksheet.cell(row_number, 4).value or "").strip()
                raw_currency = str(worksheet.cell(row_number, 5).value or "").strip().upper()
                for original_column, member_name in member_columns:
                    original_amount = _decimal(worksheet.cell(row_number, original_column).value)
                    base_amount = _decimal(worksheet.cell(row_number, original_column + 1).value)
                    if not original_amount and not base_amount:
                        continue
                    if not all(
                        (
                            member_name,
                            account_name,
                            region_name,
                            account_type_name,
                            asset_category_name,
                            raw_currency,
                        )
                    ):
                        raise AssetSnapshotWorkbookError(
                            f"{snapshot_date} 第 {row_number} 行存在金额，但基础信息不完整"
                        )
                    if raw_currency not in CURRENCY_ALIASES:
                        raise AssetSnapshotWorkbookError(
                            f"{snapshot_date} 第 {row_number} 行币种“{raw_currency}”不受支持"
                        )
                    entries.append(
                        ParsedAssetEntry(
                            source_row=row_number,
                            member_name=member_name,
                            account_name=account_name,
                            region_name=region_name,
                            account_type_name=account_type_name,
                            asset_category_name=asset_category_name,
                            currency=CURRENCY_ALIASES[raw_currency],
                            original_amount=original_amount,
                            base_amount=base_amount,
                        )
                    )
            snapshots.append(
                ParsedAssetSnapshot(
                    snapshot_date=snapshot_date,
                    usd_to_base=usd_to_base,
                    hkd_to_base=hkd_to_base,
                    entries=tuple(entries),
                )
            )
        if not snapshots:
            raise AssetSnapshotWorkbookError("文件中没有可导入的资产快照")
        return tuple(snapshots)
    finally:
        workbook.close()


def latest_snapshot_default_signature(family=None):
    snapshots = AssetBalanceSnapshot.objects.all()
    if family is not None:
        snapshots = snapshots.filter(family=family)
    latest = snapshots.order_by("-snapshot_date", "-created_at").first()
    if not latest:
        return None, ()
    signature = tuple(
        latest.entries.order_by(
            "display_order", "account__account_name", "asset_category__name"
        ).values_list(
            "member_id",
            "account_id",
            "asset_category_id",
            "currency",
            "remark",
            "display_order",
        )
    )
    return latest.pk, signature


def _verify_existing_snapshot(existing, parsed):
    if existing.usd_to_base != parsed.usd_to_base or existing.hkd_to_base != parsed.hkd_to_base:
        raise AssetSnapshotWorkbookError(
            f"{parsed.snapshot_date} 已存在，但汇率不一致："
            f"数据库 USD/HKD={existing.usd_to_base}/{existing.hkd_to_base}，"
            f"文件={parsed.usd_to_base}/{parsed.hkd_to_base}"
        )
    actual_count = existing.entries.count()
    actual_total = existing.entries.aggregate(total=Sum("base_amount"))["total"] or Decimal("0")
    if actual_count != len(parsed.entries) or actual_total.quantize(MONEY_QUANTUM) != parsed.total_base:
        raise AssetSnapshotWorkbookError(
            f"{parsed.snapshot_date} 已存在，但明细核对不一致："
            f"数据库 {actual_count} 条/{actual_total.quantize(MONEY_QUANTUM)} 元，"
            f"文件 {len(parsed.entries)} 条/{parsed.total_base} 元"
        )


def _lookup_objects(family):
    members = {member.display_name: member for member in FamilyMember.objects.filter(family=family)}
    account_types = {item.name: item for item in AccountType.objects.filter(family=family)}
    regions = {item.name: item for item in AccountRegion.objects.filter(family=family)}
    categories = {item.name: item for item in AssetCategory.objects.filter(family=family)}
    accounts = {
        (account.member.display_name, account.account_name): account
        for account in BankAccount.objects.filter(family=family).select_related("member")
    }
    return members, account_types, regions, categories, accounts


@transaction.atomic
def import_asset_snapshot_workbook(*, family, source, source_filename=None):
    parsed_snapshots = parse_asset_snapshot_workbook(source)
    latest_before = latest_snapshot_default_signature(family)
    members, account_types, regions, categories, accounts = _lookup_objects(family)
    result = AssetSnapshotImportResult()
    source_filename = source_filename or Path(getattr(source, "name", str(source))).name

    for parsed in parsed_snapshots:
        existing = (
            AssetBalanceSnapshot.objects.filter(
                family=family, snapshot_date=parsed.snapshot_date
            )
            .order_by("-created_at")
            .first()
        )
        summary = {
            "date": parsed.snapshot_date,
            "entry_count": len(parsed.entries),
            "total_base": parsed.total_base,
            "member_totals": parsed.member_totals,
        }
        result.snapshot_summaries.append(summary)
        if existing:
            _verify_existing_snapshot(existing, parsed)
            result.verified_dates.append(parsed.snapshot_date)
            continue

        snapshot = AssetBalanceSnapshot.objects.create(
            family=family,
            snapshot_date=parsed.snapshot_date,
            base_currency="CNY",
            usd_to_base=parsed.usd_to_base,
            hkd_to_base=parsed.hkd_to_base,
            remark=f"从{source_filename}导入",
            extra_data={"source_file": source_filename, "source_sheet": WORKSHEET_NAME},
        )
        entries = []
        for display_order, entry in enumerate(parsed.entries, start=1):
            member = members.get(entry.member_name)
            account_type = account_types.get(entry.account_type_name)
            region = regions.get(entry.region_name)
            category = categories.get(entry.asset_category_name)
            missing = [
                label
                for label, value in (
                    (f"成员“{entry.member_name}”", member),
                    (f"账户类型“{entry.account_type_name}”", account_type),
                    (f"账户区域“{entry.region_name}”", region),
                    (f"资产类别“{entry.asset_category_name}”", category),
                )
                if value is None
            ]
            if missing:
                raise AssetSnapshotWorkbookError(
                    f"{parsed.snapshot_date} 第 {entry.source_row} 行无法匹配：{'、'.join(missing)}"
                )
            normalized_account_name = ACCOUNT_ALIASES.get(entry.account_name, entry.account_name)
            account_key = (entry.member_name, normalized_account_name)
            account = accounts.get(account_key)
            if account is None:
                account = BankAccount.objects.create(
                    family=family,
                    member=member,
                    account_name=normalized_account_name,
                    account_type_ref=account_type,
                    account_region=region,
                    is_active=False,
                    remark="历史资产快照账户（不加入当前录入账户列表）",
                    extra_data={
                        "source_file": source_filename,
                        "source_account_name": entry.account_name,
                    },
                )
                accounts[account_key] = account
                result.historical_accounts_created.append(str(account))
            entries.append(
                AssetBalanceEntry(
                    snapshot=snapshot,
                    member=member,
                    account=account,
                    account_name=account.account_name,
                    asset_category=category,
                    currency=entry.currency,
                    original_amount=entry.original_amount,
                    base_amount=entry.base_amount,
                    display_order=display_order,
                    extra_data={
                        "source_file": source_filename,
                        "source_sheet": WORKSHEET_NAME,
                        "source_row": entry.source_row,
                    },
                )
            )
        AssetBalanceEntry.objects.bulk_create(entries)
        result.created_dates.append(parsed.snapshot_date)

    if latest_snapshot_default_signature(family) != latest_before:
        raise AssetSnapshotWorkbookError("导入改变了新增资产快照页面所依据的最新明细，已回滚")
    return result

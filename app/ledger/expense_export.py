from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8DEE8"))
MONEY_FORMAT = "#,##0.00"
DATE_FORMAT = "yyyy-mm-dd"
DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def _local_naive(value):
    if not value:
        return None
    if timezone.is_aware(value):
        return timezone.localtime(value).replace(tzinfo=None)
    return value


def build_expense_workbook(records, year):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"{year}年支出明细"
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A3"

    headers = [
        "记录ID",
        "支出日期",
        "支出时间",
        "统计开始",
        "统计结束",
        "成员",
        "支出账户",
        "一级分类",
        "二级分类",
        "三级分类",
        "金额",
        "币种",
        "商户或对象",
        "支付方式",
        "备注",
        "导入文件",
        "源文件行号",
    ]
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(headers),
    )
    title_cell = worksheet.cell(1, 1, f"{year}年家庭支出明细")
    title_cell.fill = TITLE_FILL
    title_cell.font = Font(color="FFFFFF", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 26

    for column, label in enumerate(headers, start=1):
        cell = worksheet.cell(2, column, label)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="17365D")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_index, record in enumerate(records, start=3):
        category_names = record.category.path_names if record.category else []
        values = [
            record.pk,
            record.expense_date,
            _local_naive(record.occurred_at),
            record.period_start,
            record.period_end,
            record.member.display_name,
            record.bank_account.account_name if record.bank_account else "",
            category_names[0] if category_names else "",
            category_names[1] if len(category_names) > 1 else "",
            category_names[2] if len(category_names) > 2 else "",
            float(record.amount or 0),
            record.currency,
            record.merchant,
            record.payment_method,
            record.remark,
            record.import_batch.source_filename if record.import_batch else "",
            record.import_row_number,
        ]
        for column, value in enumerate(values, start=1):
            cell = worksheet.cell(row_index, column, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="right" if column in {1, 11, 17} else "left",
                vertical="center",
            )
        for column in (2, 4, 5):
            worksheet.cell(row_index, column).number_format = DATE_FORMAT
        worksheet.cell(row_index, 3).number_format = DATETIME_FORMAT
        worksheet.cell(row_index, 11).number_format = MONEY_FORMAT

    final_row = max(2, worksheet.max_row)
    worksheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{final_row}"

    widths = [10, 13, 20, 13, 13, 12, 20, 16, 18, 18, 14, 9, 20, 14, 32, 24, 12]
    for column, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(column)].width = width

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output

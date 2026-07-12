from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db.models import Sum
from django.test import TestCase
from django.utils import timezone
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from family_core.models import AccountRegion, AccountType, AssetCategory, Family, FamilyMember

from .asset_snapshot_import import (
    AssetSnapshotWorkbookError,
    import_asset_snapshot_workbook,
    latest_snapshot_default_signature,
)
from .admin import ExpenseCategoryAdminForm, IncomeCategoryAdminForm
from .expense_import import EXPECTED_HEADERS, ExpenseWorkbookError, import_expense_workbook
from .forms import AssetBalanceEntryForm, AssetBalanceSnapshotForm, ExpenseRecordForm
from .models import (
    AnnualBudget,
    AnnualBudgetLine,
    AssetBalanceEntry,
    AssetBalanceSnapshot,
    BankAccount,
    ExpenseCategory,
    ExpenseImportBatch,
    ExpenseRecord,
    IncomeCategory,
    IncomeRecord,
)
from .views import (
    build_asset_snapshot_matrix,
    build_budget_report,
    build_cashflow_monthly_rows,
)


class LedgerExpenseImportTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="我的家庭")
        self.me = FamilyMember.objects.create(
            family=self.family,
            display_name="我",
            role=FamilyMember.ROLE_ADMIN,
        )
        self.secretary = FamilyMember.objects.create(
            family=self.family,
            display_name="孙秘书",
        )
        self.account_types = {
            name: AccountType.objects.create(family=self.family, name=name, display_order=index)
            for index, name in enumerate(("银行", "微信", "支付宝", "券商"), start=1)
        }
        self.accounts = {}
        for member in (self.me, self.secretary):
            for name in ("微信", "支付宝"):
                self.accounts[(member.display_name, name)] = BankAccount.objects.create(
                    family=self.family,
                    member=member,
                    account_name=name,
                    account_type_ref=self.account_types[name],
                )
        self.bank_account = BankAccount.objects.create(
            family=self.family,
            member=self.me,
            account_name="招商银行",
            account_type_ref=self.account_types["银行"],
        )
        self.broker_account = BankAccount.objects.create(
            family=self.family,
            member=self.me,
            account_name="测试证券",
            account_type_ref=self.account_types["券商"],
        )
        self.user = get_user_model().objects.create_user(username="tester", password="password")
        self.me.user = self.user
        self.me.save(update_fields=["user", "updated_at"])

    def workbook_upload(self, rows, filename="支出.xlsx", headers=EXPECTED_HEADERS):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "月度支出"
        worksheet.append(list(headers))
        for row in rows:
            worksheet.append(row)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        return SimpleUploadedFile(
            filename,
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def sample_rows(self):
        return [
            [datetime(2026, 6, 1, 8, 30), "我", "微信", "经常性", "餐饮", "食堂", 100, "早餐"],
            [datetime(2026, 6, 2, 9, 0), "我", "微信", "经常性", "餐饮", "", -20, "退款"],
            [datetime(2026, 6, 3, 10, 15), "孙秘书", "支付宝", "固定资产", "家居家电", "小家电", 10, "插座"],
        ]

    def test_import_creates_records_accounts_and_three_level_categories(self):
        result = import_expense_workbook(
            family=self.family,
            uploaded_file=self.workbook_upload(self.sample_rows()),
            imported_by=self.user,
        )

        self.assertFalse(result.duplicate_file)
        self.assertEqual(result.batch.row_count, 3)
        self.assertEqual(result.batch.imported_count, 3)
        self.assertEqual(result.batch.total_amount, Decimal("90"))
        self.assertEqual(ExpenseRecord.objects.count(), 3)
        refund = ExpenseRecord.objects.get(remark="退款")
        self.assertEqual(refund.amount, Decimal("-20"))
        self.assertEqual(refund.bank_account, self.accounts[("我", "微信")])
        self.assertEqual(str(refund.category), "经常性-餐饮")
        appliance = ExpenseRecord.objects.get(remark="插座")
        self.assertEqual(str(appliance.category), "固定资产-家居家电-小家电")
        self.assertEqual(timezone.localtime(appliance.occurred_at).hour, 10)

    def test_duplicate_file_and_overlapping_rows_are_not_reimported(self):
        first_upload = self.workbook_upload(self.sample_rows())
        first_bytes = first_upload.read()
        first_upload.seek(0)
        first = import_expense_workbook(family=self.family, uploaded_file=first_upload)
        duplicate = import_expense_workbook(
            family=self.family,
            uploaded_file=SimpleUploadedFile("改名.xlsx", first_bytes),
        )
        self.assertTrue(duplicate.duplicate_file)
        self.assertEqual(duplicate.batch, first.batch)
        self.assertEqual(ExpenseRecord.objects.count(), 3)

        overlap_rows = [
            self.sample_rows()[0],
            [datetime(2026, 6, 4, 12, 0), "我", "支付宝", "经常性", "交通", "", 30, "公交"],
        ]
        overlap = import_expense_workbook(
            family=self.family,
            uploaded_file=self.workbook_upload(overlap_rows, filename="次月.xlsx"),
        )
        self.assertEqual(overlap.batch.imported_count, 1)
        self.assertEqual(overlap.batch.skipped_count, 1)
        self.assertEqual(ExpenseRecord.objects.count(), 4)
        self.assertEqual(ExpenseImportBatch.objects.count(), 2)

    def test_invalid_headers_are_rejected_without_database_changes(self):
        invalid_headers = ("错误时间",) + EXPECTED_HEADERS[1:]
        with self.assertRaisesMessage(ExpenseWorkbookError, "首行表头不符合固定格式"):
            import_expense_workbook(
                family=self.family,
                uploaded_file=self.workbook_upload(self.sample_rows(), headers=invalid_headers),
            )
        self.assertFalse(ExpenseRecord.objects.exists())
        self.assertFalse(ExpenseImportBatch.objects.exists())

    def test_manual_form_uses_linked_category_dropdowns_and_allowed_accounts(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(family=self.family, name="餐饮", parent=primary)
        tertiary = ExpenseCategory.objects.create(family=self.family, name="食堂", parent=secondary)
        form = ExpenseRecordForm(
            data={
                "family": self.family.pk,
                "member": self.me.pk,
                "bank_account": self.accounts[("我", "微信")].pk,
                "primary_category": primary.pk,
                "secondary_category": secondary.pk,
                "tertiary_category": tertiary.pk,
                "expense_date": "2026-06-27",
                "amount": "25.50",
                "currency": "CNY",
                "remark": "午餐",
            }
        )
        self.assertTrue(form.is_valid(), form.errors)
        self.assertNotIn(self.broker_account, form.fields["bank_account"].queryset)
        rendered_accounts = str(form["bank_account"])
        self.assertIn("data-member-id", rendered_accounts)
        self.assertIn("我 - 微信", rendered_accounts)
        self.assertNotIn("测试证券", rendered_accounts)
        record = form.save()
        self.assertEqual(record.category, tertiary)
        self.assertEqual(record.bank_account, self.accounts[("我", "微信")])
        self.assertEqual(record.period_start.isoformat(), "2026-06-27")
        self.client.force_login(self.user)
        response = self.client.get(reverse("ledger:expense_year_detail", args=[2026]))
        self.assertContains(response, "<th>支出账户</th>", html=True)
        self.assertContains(response, "<td>微信</td>", html=True)

    def test_manual_form_rejects_account_owned_by_another_member(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(family=self.family, name="餐饮", parent=primary)
        form = ExpenseRecordForm(
            data={
                "family": self.family.pk,
                "member": self.me.pk,
                "bank_account": self.accounts[("孙秘书", "微信")].pk,
                "primary_category": primary.pk,
                "secondary_category": secondary.pk,
                "tertiary_category": "",
                "expense_date": "2026-06-27",
                "amount": "5",
                "currency": "CNY",
                "remark": "",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("bank_account", form.errors)

    def test_monthly_cashflow_summary_includes_family_totals(self):
        record_date = date(2026, 6, 15)
        IncomeRecord.objects.create(
            family=self.family,
            member=self.me,
            income_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("100"),
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.me,
            bank_account=self.accounts[("我", "微信")],
            expense_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("30"),
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.secretary,
            bank_account=self.accounts[("孙秘书", "支付宝")],
            expense_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("20"),
        )

        _, sections = build_cashflow_monthly_rows(2026)
        june = next(row for row in sections[0]["rows"] if row["month"] == 6)
        self.assertEqual(june["income_total"], Decimal("100"))
        self.assertEqual(june["expense_total"], Decimal("50"))
        self.assertEqual(june["net_total"], Decimal("50"))
        self.assertEqual(sections[0]["total"]["expense_total"], Decimal("50"))

        self.client.force_login(self.user)
        response = self.client.get(reverse("ledger:cashflow_summary_year", args=[2026]))
        self.assertContains(response, "<th>家庭合计</th>", count=3, html=True)

    def test_monthly_summary_reuses_linked_expense_pies_with_month_filter(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )
        tertiary = ExpenseCategory.objects.create(
            family=self.family,
            name="食堂",
            parent=secondary,
        )
        for record_date, amount in (
            (date(2026, 6, 20), Decimal("10000")),
            (date(2026, 7, 20), Decimal("20000")),
        ):
            ExpenseRecord.objects.create(
                family=self.family,
                member=self.me,
                bank_account=self.accounts[("我", "微信")],
                category=tertiary,
                expense_date=record_date,
                period_start=record_date,
                period_end=record_date,
                amount=amount,
            )
        self.client.force_login(self.user)

        june_response = self.client.get(
            reverse("ledger:cashflow_summary_year", args=[2026]),
            {"category_month": "6"},
        )

        self.assertEqual(june_response.status_code, 200)
        self.assertEqual(june_response.context["selected_category_month"], 6)
        june_pies = june_response.context["expense_category_pie_data"]
        self.assertEqual(june_pies["unit"], "元")
        self.assertEqual(sum(item["value"] for item in june_pies["primary"]), 10000.0)
        self.assertEqual(sum(item["value"] for item in june_pies["secondary"]), 10000.0)
        self.assertEqual(sum(item["value"] for item in june_pies["tertiary"]), 10000.0)
        self.assertContains(june_response, 'id="monthly-expense-category-pies"')
        self.assertContains(june_response, 'name="category_month"')
        self.assertContains(june_response, 'value="12"')
        self.assertContains(june_response, "js/expense_category_pies.js")
        self.assertContains(
            june_response,
            'action="/ledger/expenses/2026/summary/#monthly-expense-category-pies"',
        )
        rendered_html = june_response.content.decode()
        self.assertLess(
            rendered_html.index("</table>"),
            rendered_html.index('id="monthly-expense-category-pies"'),
        )

        annual_response = self.client.get(
            reverse("ledger:cashflow_summary_year", args=[2026]),
            {"category_month": "all"},
        )
        annual_pies = annual_response.context["expense_category_pie_data"]
        self.assertEqual(annual_pies["unit"], "元")
        self.assertEqual(sum(item["value"] for item in annual_pies["primary"]), 30000.0)
        self.assertEqual(sum(item["value"] for item in annual_pies["secondary"]), 30000.0)
        self.assertEqual(sum(item["value"] for item in annual_pies["tertiary"]), 30000.0)

    def test_year_expense_export_contains_every_record_and_category_level(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )
        tertiary = ExpenseCategory.objects.create(
            family=self.family,
            name="食堂",
            parent=secondary,
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.me,
            bank_account=self.accounts[("我", "微信")],
            category=tertiary,
            expense_date=date(2026, 6, 20),
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            amount=Decimal("25.50"),
            merchant="家庭食堂",
            remark="午餐",
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.me,
            expense_date=date(2025, 12, 20),
            amount=Decimal("99"),
            remark="其他年份",
        )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("ledger:expense_year_export", args=[2026])
        )
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        worksheet = workbook["2026年支出明细"]

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual(worksheet["A2"].value, "记录ID")
        self.assertEqual(worksheet["H3"].value, "经常性")
        self.assertEqual(worksheet["I3"].value, "餐饮")
        self.assertEqual(worksheet["J3"].value, "食堂")
        self.assertEqual(worksheet["K3"].value, 25.5)
        self.assertEqual(worksheet["K3"].number_format, "#,##0.00")
        self.assertEqual(worksheet["O3"].value, "午餐")
        workbook.close()

    def test_expense_index_cashflow_chart_supports_single_year_and_all_years(self):
        recurring = ExpenseCategory.objects.create(family=self.family, name="经常性")
        dining = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=recurring,
        )
        cafeteria = ExpenseCategory.objects.create(
            family=self.family,
            name="食堂",
            parent=dining,
        )
        fixed_asset = ExpenseCategory.objects.create(family=self.family, name="固定资产")
        home_appliance = ExpenseCategory.objects.create(
            family=self.family,
            name="家居家电",
            parent=fixed_asset,
        )
        small_appliance = ExpenseCategory.objects.create(
            family=self.family,
            name="小家电",
            parent=home_appliance,
        )
        for record_date, income_amount, expense_amount, category in (
            (date(2025, 12, 15), Decimal("20000"), Decimal("5000"), cafeteria),
            (date(2026, 1, 15), Decimal("10000"), Decimal("12000"), cafeteria),
            (date(2026, 6, 15), Decimal("30000"), Decimal("5000"), small_appliance),
        ):
            IncomeRecord.objects.create(
                family=self.family,
                member=self.me,
                income_date=record_date,
                period_start=record_date,
                period_end=record_date,
                amount=income_amount,
            )
            ExpenseRecord.objects.create(
                family=self.family,
                member=self.me,
                bank_account=self.accounts[("我", "微信")],
                category=category,
                expense_date=record_date,
                period_start=record_date,
                period_end=record_date,
                amount=expense_amount,
            )
        self.client.force_login(self.user)

        yearly_response = self.client.get(
            reverse("ledger:expense_list"),
            {"trend_year": "2026", "category_year": "2026"},
        )
        self.assertEqual(yearly_response.status_code, 200)
        yearly_trend = yearly_response.context["cashflow_trend_data"]
        self.assertEqual(yearly_trend["mode"], "monthly")
        self.assertEqual(yearly_trend["labels"], [f"{month}月" for month in range(1, 13)])
        self.assertEqual(yearly_trend["income"][0], 1.0)
        self.assertEqual(yearly_trend["expense"][0], 1.2)
        self.assertEqual(yearly_trend["net"][0], -0.2)
        self.assertEqual(yearly_trend["income"][5], 3.0)
        self.assertContains(yearly_response, 'id="cashflow-trend-chart"')
        self.assertContains(yearly_response, 'value="all"')
        self.assertContains(yearly_response, "js/cashflow_trend.js")
        self.assertContains(yearly_response, "js/expense_category_pies.js")
        self.assertContains(yearly_response, 'id="primary-expense-pie"')
        category_pies = yearly_response.context["expense_category_pie_data"]
        primary_values = {
            item["name"]: item["value"]
            for item in category_pies["primary"]
        }
        self.assertEqual(primary_values, {"经常性": 1.2, "固定资产": 0.5})
        recurring_secondary = next(
            item for item in category_pies["secondary"] if item["name"] == "餐饮"
        )
        self.assertEqual(recurring_secondary["parent_id"], recurring.pk)
        recurring_tertiary = next(
            item for item in category_pies["tertiary"] if item["name"] == "食堂"
        )
        self.assertEqual(recurring_tertiary["parent_id"], dining.pk)
        self.assertEqual(recurring_tertiary["primary_id"], recurring.pk)

        all_response = self.client.get(
            reverse("ledger:expense_list"),
            {"trend_year": "all", "category_year": "all"},
        )
        all_trend = all_response.context["cashflow_trend_data"]
        self.assertEqual(all_trend["mode"], "yearly")
        self.assertEqual(all_trend["labels"], ["2025", "2026"])
        self.assertEqual(all_trend["income"], [2.0, 4.0])
        self.assertEqual(all_trend["expense"], [0.5, 1.7])
        self.assertEqual(all_trend["net"], [1.5, 2.3])
        all_primary_values = {
            item["name"]: item["value"]
            for item in all_response.context["expense_category_pie_data"]["primary"]
        }
        self.assertEqual(all_primary_values, {"经常性": 1.7, "固定资产": 0.5})

    def test_expense_pie_levels_reconcile_direct_and_nested_classifications(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )
        tertiary = ExpenseCategory.objects.create(
            family=self.family,
            name="食堂",
            parent=secondary,
        )
        record_date = date(2026, 6, 20)
        for category, amount in (
            (primary, Decimal("10000")),
            (secondary, Decimal("20000")),
            (tertiary, Decimal("30000")),
        ):
            ExpenseRecord.objects.create(
                family=self.family,
                member=self.me,
                bank_account=self.accounts[("我", "微信")],
                category=category,
                expense_date=record_date,
                period_start=record_date,
                period_end=record_date,
                amount=amount,
            )
        self.client.force_login(self.user)

        response = self.client.get(
            reverse("ledger:expense_list"),
            {"category_year": "2026"},
        )

        self.assertEqual(response.status_code, 200)
        pie_data = response.context["expense_category_pie_data"]
        primary_total = sum(item["value"] for item in pie_data["primary"])
        secondary_total = sum(item["value"] for item in pie_data["secondary"])
        tertiary_total = sum(item["value"] for item in pie_data["tertiary"])
        self.assertEqual(primary_total, 6.0)
        self.assertEqual(secondary_total, primary_total)
        self.assertEqual(tertiary_total, primary_total)

        primary_value = next(
            item["value"] for item in pie_data["primary"] if item["id"] == primary.pk
        )
        primary_secondary_items = [
            item for item in pie_data["secondary"] if item["parent_id"] == primary.pk
        ]
        primary_tertiary_items = [
            item for item in pie_data["tertiary"] if item["primary_id"] == primary.pk
        ]
        self.assertEqual(sum(item["value"] for item in primary_secondary_items), primary_value)
        self.assertEqual(sum(item["value"] for item in primary_tertiary_items), primary_value)

        real_secondary_value = next(
            item["value"] for item in primary_secondary_items if item["id"] == secondary.pk
        )
        secondary_tertiary_total = sum(
            item["value"]
            for item in primary_tertiary_items
            if item["parent_id"] == secondary.pk
        )
        self.assertEqual(secondary_tertiary_total, real_secondary_value)
        self.assertTrue(
            any(item["name"] == "未细分至二级" for item in primary_secondary_items)
        )
        self.assertTrue(
            any(item["name"] == "未细分至三级" for item in primary_tertiary_items)
        )
        self.assertContains(
            response,
            'action="/ledger/expenses/#expense-category-pies"',
        )

    def test_month_expense_filters_recalculate_member_and_family_totals(self):
        primary = ExpenseCategory.objects.create(family=self.family, name="经常性")
        secondary = ExpenseCategory.objects.create(family=self.family, name="餐饮", parent=primary)
        cafeteria = ExpenseCategory.objects.create(family=self.family, name="食堂", parent=secondary)
        takeaway = ExpenseCategory.objects.create(family=self.family, name="外卖", parent=secondary)
        record_date = date(2026, 6, 20)
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.me,
            bank_account=self.accounts[("我", "微信")],
            category=cafeteria,
            expense_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("25.50"),
            remark="筛选目标",
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.me,
            bank_account=self.accounts[("我", "支付宝")],
            category=takeaway,
            expense_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("30"),
            remark="同成员不同账户",
        )
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.secretary,
            bank_account=self.accounts[("孙秘书", "支付宝")],
            category=cafeteria,
            expense_date=record_date,
            period_start=record_date,
            period_end=record_date,
            amount=Decimal("40"),
            remark="不同成员",
        )

        self.client.force_login(self.user)
        response = self.client.get(
            reverse("ledger:expense_month_detail", args=[2026, 6]),
            {
                "member": self.me.pk,
                "bank_account": self.accounts[("我", "微信")].pk,
                "primary_category": primary.pk,
                "secondary_category": secondary.pk,
                "tertiary_category": cafeteria.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["expense_rows"]), 1)
        self.assertEqual(response.context["expense_rows"][0]["remark"], "筛选目标")
        totals = {
            row["member"].display_name: row["amount"]
            for row in response.context["expense_member_totals"]
        }
        self.assertEqual(totals["我"], Decimal("25.50"))
        self.assertEqual(totals["孙秘书"], Decimal("0"))
        self.assertEqual(response.context["expense_family_total"], Decimal("25.50"))
        self.assertContains(response, "已按条件筛选，合计已重新计算")
        self.assertContains(response, 'aria-label="按三级分类筛选"')
        self.assertContains(response, "expense-filter-scroll:")
        self.assertContains(response, "sessionStorage.setItem")
        self.assertContains(response, "window.scrollTo(0, savedScrollPosition)")

        year_response = self.client.get(
            reverse("ledger:expense_year_detail", args=[2026]),
            {
                "member": self.me.pk,
                "bank_account": self.accounts[("我", "微信")].pk,
                "primary_category": primary.pk,
                "secondary_category": secondary.pk,
                "tertiary_category": cafeteria.pk,
            },
        )
        self.assertEqual(year_response.status_code, 200)
        self.assertEqual(len(year_response.context["expense_rows"]), 1)
        self.assertEqual(year_response.context["expense_family_total"], Decimal("25.50"))
        self.assertContains(year_response, 'aria-label="按成员筛选"')
        self.assertContains(year_response, "已按条件筛选，合计已重新计算")


class AssetSnapshotImportTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="我的家庭")
        self.me = FamilyMember.objects.create(
            family=self.family,
            display_name="我",
            role=FamilyMember.ROLE_ADMIN,
        )
        self.secretary = FamilyMember.objects.create(
            family=self.family,
            display_name="孙秘书",
        )
        self.bank_type = AccountType.objects.create(family=self.family, name="银行")
        self.broker_type = AccountType.objects.create(family=self.family, name="券商")
        self.domestic = AccountRegion.objects.create(family=self.family, name="境内")
        self.overseas = AccountRegion.objects.create(family=self.family, name="境外")
        self.cash = AssetCategory.objects.create(
            family=self.family, name="现金及现金等价物", code="cash"
        )
        self.fund = AssetCategory.objects.create(
            family=self.family, name="基金类", code="fund"
        )
        self.current_account = BankAccount.objects.create(
            family=self.family,
            member=self.me,
            account_name="当前账户",
            account_type_ref=self.bank_type,
            account_region=self.domestic,
        )
        self.latest = AssetBalanceSnapshot.objects.create(
            family=self.family,
            snapshot_date=date(2026, 5, 31),
            usd_to_base=Decimal("6.78730000"),
            hkd_to_base=Decimal("0.86560000"),
        )
        AssetBalanceEntry.objects.create(
            snapshot=self.latest,
            member=self.me,
            account=self.current_account,
            account_name=self.current_account.account_name,
            asset_category=self.cash,
            currency="CNY",
            original_amount=Decimal("100"),
            base_amount=Decimal("100"),
            display_order=1,
            remark="保留此模板",
        )

    def workbook_source(self, existing_total=100):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "账户余额NEW"
        worksheet.cell(1, 6, datetime(2024, 12, 31))
        worksheet.cell(2, 7, 7.3)
        worksheet.cell(2, 9, 0.94)
        worksheet.cell(3, 6, "我")
        worksheet.cell(3, 8, "孙秘书")
        worksheet.cell(1, 10, datetime(2026, 5, 31))
        worksheet.cell(2, 11, 6.7873)
        worksheet.cell(2, 13, 0.8656)
        worksheet.cell(3, 10, "我")
        worksheet.cell(3, 12, "孙秘书")
        for column, value in enumerate(
            ("账户名称", "账户地区", "账户类型", "资产类别", "币种"),
            start=1,
        ):
            worksheet.cell(4, column, value)
        historical_rows = (
            ("历史银行", "境内", "银行", "现金", "RMB", 50, 50, 0, 0),
            ("境外券商", "境外", "券商", "股指基金", "USD", 10, 73, 0, 0),
        )
        for row_number, row in enumerate(historical_rows, start=5):
            for column, value in enumerate(row, start=1):
                worksheet.cell(row_number, column, value)
        worksheet.cell(7, 1, "当前账户")
        worksheet.cell(7, 2, "境内")
        worksheet.cell(7, 3, "银行")
        worksheet.cell(7, 4, "现金")
        worksheet.cell(7, 5, "RMB")
        worksheet.cell(7, 10, existing_total)
        worksheet.cell(7, 11, existing_total)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)
        stream.name = "资产快照录入.xlsx"
        return stream

    def test_imports_older_snapshot_and_preserves_latest_default_signature(self):
        before = latest_snapshot_default_signature()
        result = import_asset_snapshot_workbook(
            family=self.family,
            source=self.workbook_source(),
            source_filename="资产快照录入.xlsx",
        )

        self.assertEqual(result.created_dates, [date(2024, 12, 31)])
        self.assertEqual(result.verified_dates, [date(2026, 5, 31)])
        imported = AssetBalanceSnapshot.objects.get(snapshot_date=date(2024, 12, 31))
        self.assertEqual(imported.entries.count(), 2)
        self.assertEqual(
            imported.entries.aggregate(total=Sum("base_amount"))["total"],
            Decimal("123"),
        )
        self.assertFalse(
            BankAccount.objects.get(member=self.me, account_name="历史银行").is_active
        )
        self.assertFalse(
            BankAccount.objects.get(member=self.me, account_name="境外券商").is_active
        )
        self.assertEqual(latest_snapshot_default_signature(), before)

    def test_existing_snapshot_mismatch_rolls_back_every_change(self):
        before = latest_snapshot_default_signature()
        with self.assertRaises(AssetSnapshotWorkbookError):
            import_asset_snapshot_workbook(
                family=self.family,
                source=self.workbook_source(existing_total=99),
                source_filename="资产快照录入.xlsx",
            )

        self.assertEqual(AssetBalanceSnapshot.objects.count(), 1)
        self.assertFalse(BankAccount.objects.filter(account_name="历史银行").exists())
        self.assertEqual(latest_snapshot_default_signature(), before)

    def test_asset_snapshot_list_builds_monthly_and_yearly_trend_in_ten_thousands(self):
        older = AssetBalanceSnapshot.objects.create(
            family=self.family,
            snapshot_date=date(2025, 12, 31),
            usd_to_base=Decimal("7"),
            hkd_to_base=Decimal("0.9"),
        )
        AssetBalanceEntry.objects.create(
            snapshot=older,
            member=self.me,
            account=self.current_account,
            account_name=self.current_account.account_name,
            asset_category=self.cash,
            currency="CNY",
            original_amount=Decimal("80000"),
            base_amount=Decimal("80000"),
            display_order=1,
        )
        AssetBalanceEntry.objects.create(
            snapshot=older,
            member=self.secretary,
            account_name="历史账户",
            asset_category=self.cash,
            currency="CNY",
            original_amount=Decimal("20000"),
            base_amount=Decimal("20000"),
            display_order=2,
        )
        user = get_user_model().objects.create_user(
            username="trend-tester",
            password="password",
        )
        self.me.user = user
        self.me.save(update_fields=["user", "updated_at"])
        self.client.force_login(user)

        response = self.client.get(reverse("ledger:asset_snapshot_list"))

        self.assertEqual(response.status_code, 200)
        trend_data = response.context["trend_data"]
        self.assertEqual(trend_data["monthly"]["labels"], ["2025-12", "2026-05"])
        self.assertEqual(trend_data["yearly"]["labels"], ["2025", "2026"])
        monthly_series = {
            series["name"]: series["values"]
            for series in trend_data["monthly"]["series"]
        }
        self.assertEqual(monthly_series["我"], [8.0, 0.01])
        self.assertEqual(monthly_series["孙秘书"], [2.0, 0.0])
        self.assertEqual(monthly_series["家庭合计"], [10.0, 0.01])
        self.assertContains(response, 'id="asset-trend-chart"')
        self.assertContains(response, 'data-trend-period="monthly"')
        self.assertContains(response, 'data-trend-period="yearly"')
        self.assertContains(response, "js/asset_snapshot_trend.js")
        rendered_html = response.content.decode()
        self.assertLess(
            rendered_html.index("<h1>资产快照</h1>"),
            rendered_html.index("<h1>资产余额趋势</h1>"),
        )
        self.assertNotIn("asset-trend-summary", rendered_html)
        self.assertIn('class="asset-trend-empty" hidden', rendered_html)


class AssetSnapshotWorkspaceTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="我的家庭")
        self.member = FamilyMember.objects.create(
            family=self.family,
            display_name="我",
        )
        self.secretary = FamilyMember.objects.create(
            family=self.family,
            display_name="孙秘书",
        )
        self.category = AssetCategory.objects.create(
            family=self.family,
            name="现金",
        )
        self.account = BankAccount.objects.create(
            family=self.family,
            member=self.member,
            account_name="测试账户一",
        )
        self.second_account = BankAccount.objects.create(
            family=self.family,
            member=self.member,
            account_name="测试账户二",
        )
        self.user = get_user_model().objects.create_user(
            username="snapshot-workspace-tester",
            password="password",
        )
        self.member.user = self.user
        self.member.save(update_fields=["user", "updated_at"])
        self.client.force_login(self.user)

    def test_decimal_inputs_render_with_two_decimal_places(self):
        snapshot = AssetBalanceSnapshot(
            family=self.family,
            snapshot_date=date(2026, 6, 30),
            usd_to_base=Decimal("7.12345678"),
            hkd_to_base=Decimal("0.91234567"),
        )
        snapshot_form = AssetBalanceSnapshotForm(instance=snapshot)
        entry = AssetBalanceEntry(
            snapshot=snapshot,
            member=self.member,
            account=self.account,
            asset_category=self.category,
            original_amount=Decimal("123.4567"),
        )
        entry_form = AssetBalanceEntryForm(instance=entry)

        self.assertIn('value="7.12345678"', str(snapshot_form["usd_to_base"]))
        self.assertIn('value="0.91234567"', str(snapshot_form["hkd_to_base"]))
        self.assertIn('value="123.46"', str(entry_form["original_amount"]))

    def test_save_draft_and_persist_manual_row_order(self):
        response = self.client.post(
            reverse("ledger:asset_snapshot_create"),
            {
                "family": self.family.pk,
                "snapshot_date": "2026-06-30",
                "base_currency": "CNY",
                "usd_to_base": "7.20",
                "hkd_to_base": "0.92",
                "remark": "草稿测试",
                "save_action": "draft",
                "entries-TOTAL_FORMS": "2",
                "entries-INITIAL_FORMS": "0",
                "entries-MIN_NUM_FORMS": "0",
                "entries-MAX_NUM_FORMS": "1000",
                "entries-0-member": self.member.pk,
                "entries-0-account": self.account.pk,
                "entries-0-asset_category": self.category.pk,
                "entries-0-currency": "CNY",
                "entries-0-original_amount": "",
                "entries-0-display_order": "2",
                "entries-0-remark": "",
                "entries-1-member": self.member.pk,
                "entries-1-account": self.second_account.pk,
                "entries-1-asset_category": self.category.pk,
                "entries-1-currency": "CNY",
                "entries-1-original_amount": "20.00",
                "entries-1-display_order": "1",
                "entries-1-remark": "",
            },
        )

        snapshot = AssetBalanceSnapshot.objects.get()
        self.assertRedirects(
            response,
            reverse("ledger:asset_snapshot_detail", args=[snapshot.pk]),
        )
        self.assertTrue(snapshot.is_draft)
        self.assertEqual(
            list(
                snapshot.entries.order_by("display_order").values_list(
                    "account_id",
                    "display_order",
                )
            ),
            [(self.second_account.pk, 1), (self.account.pk, 2)],
        )
        self.assertEqual(
            snapshot.entries.get(account=self.account).original_amount,
            Decimal("0"),
        )

    def test_exchange_gain_combines_usd_and_hkd(self):
        AssetBalanceSnapshot.objects.create(
            family=self.family,
            snapshot_date=date(2026, 5, 31),
            usd_to_base=Decimal("7.00"),
            hkd_to_base=Decimal("0.90"),
        )
        current = AssetBalanceSnapshot.objects.create(
            family=self.family,
            snapshot_date=date(2026, 6, 30),
            usd_to_base=Decimal("7.20"),
            hkd_to_base=Decimal("0.95"),
        )
        AssetBalanceEntry.objects.create(
            snapshot=current,
            member=self.member,
            account=self.account,
            account_name=self.account.account_name,
            asset_category=self.category,
            currency="USD",
            original_amount=Decimal("100"),
            base_amount=Decimal("720"),
            display_order=1,
        )
        AssetBalanceEntry.objects.create(
            snapshot=current,
            member=self.member,
            account=self.second_account,
            account_name=self.second_account.account_name,
            asset_category=self.category,
            currency="HKD",
            original_amount=Decimal("1000"),
            base_amount=Decimal("950"),
            display_order=2,
        )

        matrix = build_asset_snapshot_matrix(current)
        exchange_gain_row = matrix[4]

        self.assertEqual(exchange_gain_row["cells"], [Decimal("70.00")])
        self.assertEqual(exchange_gain_row["total"], Decimal("70.00"))

    def test_export_contains_all_snapshot_blocks_and_two_decimal_format(self):
        snapshot = AssetBalanceSnapshot.objects.create(
            family=self.family,
            snapshot_date=date(2026, 6, 30),
            usd_to_base=Decimal("7.20"),
            hkd_to_base=Decimal("0.92"),
        )
        AssetBalanceEntry.objects.create(
            snapshot=snapshot,
            member=self.member,
            account=self.account,
            account_name=self.account.account_name,
            asset_category=self.category,
            currency="CNY",
            original_amount=Decimal("123.45"),
            base_amount=Decimal("123.45"),
            display_order=1,
        )

        response = self.client.get(reverse("ledger:asset_snapshot_export"))
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        worksheet = workbook["资产快照"]
        values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        ]

        self.assertEqual(response.status_code, 200)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn("2026年06月30日 资产快照", values)
        self.assertIn("汇兑损益金额", values)
        self.assertEqual(worksheet["D3"].number_format, "#,##0.00")
        workbook.close()


class LedgerOverviewChartTests(TestCase):
    def test_overview_uses_latest_assets_current_year_budget_and_investment_returns(self):
        today = timezone.localdate()
        family = Family.objects.create(name="我的家庭")
        me = FamilyMember.objects.create(family=family, display_name="我")
        secretary = FamilyMember.objects.create(family=family, display_name="孙秘书")
        cash = AssetCategory.objects.create(family=family, name="现金")
        fund = AssetCategory.objects.create(family=family, name="基金")
        income_category = IncomeCategory.objects.create(family=family, name="工资")
        expense_category = ExpenseCategory.objects.create(family=family, name="生活")

        previous = AssetBalanceSnapshot.objects.create(
            family=family,
            snapshot_date=date(today.year - 1, 12, 31),
        )
        latest = AssetBalanceSnapshot.objects.create(
            family=family,
            snapshot_date=today,
        )
        for snapshot, member, category, amount in (
            (previous, me, cash, "100"),
            (previous, secretary, cash, "200"),
            (latest, me, cash, "180"),
            (latest, me, fund, "20"),
            (latest, secretary, cash, "250"),
        ):
            AssetBalanceEntry.objects.create(
                snapshot=snapshot,
                member=member,
                asset_category=category,
                base_amount=Decimal(amount),
                original_amount=Decimal(amount),
            )

        IncomeRecord.objects.create(
            family=family,
            member=me,
            category=income_category,
            income_date=today,
            amount=Decimal("30"),
        )
        IncomeRecord.objects.create(
            family=family,
            member=secretary,
            category=income_category,
            income_date=today,
            amount=Decimal("10"),
        )
        ExpenseRecord.objects.create(
            family=family,
            member=me,
            category=expense_category,
            expense_date=today,
            amount=Decimal("5"),
        )
        ExpenseRecord.objects.create(
            family=family,
            member=secretary,
            category=expense_category,
            expense_date=today,
            amount=Decimal("5"),
        )
        budget = AnnualBudget.objects.create(family=family, year=today.year)
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_INCOME,
            income_category=income_category,
            annual_amount=Decimal("1000"),
        )
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            expense_category=expense_category,
            annual_amount=Decimal("400"),
        )
        user = get_user_model().objects.create_user(
            username="overview-chart-tester",
            password="password",
        )
        me.user = user
        me.save(update_fields=["user", "updated_at"])
        self.client.force_login(user)

        response = self.client.get(reverse("ledger:overview"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["year_income"], Decimal("40"))
        self.assertEqual(response.context["year_expense"], Decimal("10"))
        asset_charts = response.context["chart_data"]["asset_charts"]
        self.assertEqual([chart["label"] for chart in asset_charts], ["家庭合计", "我", "孙秘书"])
        self.assertEqual(
            {item["name"]: item["value"] for item in asset_charts[0]["items"]},
            {"基金": 20.0, "现金": 430.0},
        )
        budget_items = response.context["chart_data"]["budget"]["items"]
        self.assertEqual(
            [item["value"] for item in budget_items],
            [1000.0, 40.0, 400.0, 10.0, 600.0, 30.0],
        )
        self.assertEqual(
            response.context["chart_data"]["investment_returns"],
            [
                {"name": "家庭合计", "value": 120.0},
                {"name": "我", "value": 75.0},
                {"name": "孙秘书", "value": 45.0},
            ],
        )
        self.assertContains(response, "本年收入")
        self.assertNotContains(response, "本月收入")
        self.assertContains(response, "js/ledger_overview.js")

        budget_response = self.client.get(
            reverse("ledger:annual_budget_detail", args=[budget.pk])
        )
        self.assertNotContains(budget_response, "月度节奏对比")
        self.assertNotContains(budget_response, "budget-month-table")


class AnnualBudgetReportTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="我的家庭")
        self.member = FamilyMember.objects.create(
            family=self.family,
            display_name="我",
        )
        self.user = get_user_model().objects.create_user(
            username="budget-report-tester",
            password="password",
        )
        self.member.user = self.user
        self.member.save(update_fields=["user", "updated_at"])
        self.client.force_login(self.user)

    def test_legacy_budget_category_collects_current_descendant_records(self):
        legacy_root = ExpenseCategory.objects.create(
            family=self.family,
            name="经营性",
        )
        legacy_food = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=legacy_root,
        )
        current_root = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性",
        )
        current_food = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=current_root,
        )
        takeaway = ExpenseCategory.objects.create(
            family=self.family,
            name="外卖",
            parent=current_food,
        )
        budget = AnnualBudget.objects.create(family=self.family, year=2026)
        line = AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            expense_category=legacy_food,
            annual_amount=Decimal("12000"),
            extra_data={"category_path": "经营性-餐饮"},
        )
        legacy_food.delete()
        line.refresh_from_db()
        self.assertIsNone(line.expense_category)
        ExpenseRecord.objects.create(
            family=self.family,
            member=self.member,
            category=takeaway,
            expense_date=date(2026, 6, 30),
            period_start=date(2026, 6, 1),
            period_end=date(2026, 6, 30),
            amount=Decimal("88.50"),
        )

        report = build_budget_report(budget)

        self.assertEqual(report["expense_line_rows"][0]["actual"], Decimal("88.50"))

    def test_budget_list_includes_savings_plan(self):
        budget = AnnualBudget.objects.create(family=self.family, year=2026)
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_INCOME,
            annual_amount=Decimal("100"),
        )
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            annual_amount=Decimal("40"),
        )

        response = self.client.get(reverse("ledger:annual_budget_list"))

        self.assertContains(response, "攒钱计划")
        self.assertEqual(response.context["rows"][0]["savings_budget"], Decimal("60"))

    def test_expense_summary_rows_follow_current_primary_categories(self):
        recurring = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性",
        )
        recurring_child = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=recurring,
        )
        non_recurring = ExpenseCategory.objects.create(
            family=self.family,
            name="非经常性",
        )
        non_recurring_child = ExpenseCategory.objects.create(
            family=self.family,
            name="大额支出",
            parent=non_recurring,
        )
        old_root = ExpenseCategory.objects.create(
            family=self.family,
            name="经营性",
        )
        old_child = ExpenseCategory.objects.create(
            family=self.family,
            name="旧分类",
            parent=old_root,
        )
        budget = AnnualBudget.objects.create(family=self.family, year=2026)
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            expense_category=recurring_child,
            annual_amount=Decimal("100"),
        )
        AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            expense_category=non_recurring_child,
            annual_amount=Decimal("200"),
        )
        old_line = AnnualBudgetLine.objects.create(
            budget=budget,
            line_type=AnnualBudgetLine.LINE_TYPE_EXPENSE,
            expense_category=old_child,
            annual_amount=Decimal("300"),
            extra_data={"category_path": "经营性-旧分类"},
        )
        old_child.delete()
        old_root.delete()
        old_line.refresh_from_db()

        report = build_budget_report(budget)
        summary_rows = {
            row["label"]: row for row in report["expense_summary_rows"]
        }

        self.assertEqual(
            set(summary_rows),
            {"经常性汇总", "非经常性汇总", "支出汇总"},
        )
        self.assertEqual(summary_rows["经常性汇总"]["budget"], Decimal("100"))
        self.assertEqual(summary_rows["非经常性汇总"]["budget"], Decimal("200"))
        self.assertNotIn("经营性汇总", summary_rows)


class CategoryManagementTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="我的家庭")
        self.user = get_user_model().objects.create_user(
            username="category-tester",
            password="password",
        )
        FamilyMember.objects.create(
            family=self.family,
            user=self.user,
            display_name="分类维护测试成员",
        )

    def test_expense_admin_form_creates_explicit_three_level_path(self):
        primary = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性支出",
        )
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )
        form = ExpenseCategoryAdminForm(
            data={
                "family": self.family.pk,
                "category_level": "3",
                "primary_category": primary.pk,
                "secondary_category": secondary.pk,
                "name": "工作餐",
                "is_active": "on",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        category = form.save()
        self.assertEqual(category.parent, secondary)
        self.assertEqual(category.category_level, 3)
        self.assertEqual(str(category), "经常性支出-餐饮-工作餐")

    def test_admin_form_rejects_secondary_from_another_primary(self):
        first_primary = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性支出",
        )
        second_primary = ExpenseCategory.objects.create(
            family=self.family,
            name="固定资产",
        )
        wrong_secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="家居家电",
            parent=second_primary,
        )
        form = ExpenseCategoryAdminForm(
            data={
                "family": self.family.pk,
                "category_level": "3",
                "primary_category": first_primary.pk,
                "secondary_category": wrong_secondary.pk,
                "name": "小家电",
                "is_active": "on",
            }
        )

        self.assertFalse(form.is_valid())
        self.assertIn("secondary_category", form.errors)

    def test_income_admin_form_uses_same_level_structure(self):
        primary = IncomeCategory.objects.create(family=self.family, name="经营收入")
        form = IncomeCategoryAdminForm(
            data={
                "family": self.family.pk,
                "category_level": "2",
                "primary_category": primary.pk,
                "secondary_category": "",
                "name": "服务收入",
                "is_active": "on",
            }
        )

        self.assertTrue(form.is_valid(), form.errors)
        category = form.save()
        self.assertEqual(str(category), "经营收入-服务收入")
        self.assertEqual(category.category_level, 2)

    def test_front_page_is_read_only_and_legacy_add_url_redirects_to_admin(self):
        ExpenseCategory.objects.create(family=self.family, name="经常性支出")
        self.client.force_login(self.user)

        response = self.client.get(reverse("ledger:category_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "收入分类统一在后台按一级、二级维护")
        self.assertContains(response, "支出分类统一在后台按一级、二级、三级维护")
        self.assertContains(response, 'class="basics-table-scroll"', count=3)
        self.assertNotContains(response, "新增支出分类")
        self.assertNotContains(response, ">编辑</a>")

        redirect_response = self.client.get(reverse("ledger:expense_category_create"))
        self.assertRedirects(
            redirect_response,
            reverse("admin:ledger_expensecategory_add"),
            fetch_redirect_response=False,
        )

    def test_ledger_overview_does_not_show_direct_create_buttons(self):
        self.client.force_login(self.user)

        response = self.client.get(reverse("ledger:overview"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, ">新增收入</a>")
        self.assertNotContains(response, ">新增支出</a>")
        self.assertNotContains(response, ">新增资产快照</a>")

    def test_admin_add_pages_expose_correct_level_fields_instead_of_parent(self):
        primary = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性支出",
        )
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )
        staff = get_user_model().objects.create_superuser(
            username="category-admin",
            password="password",
            email="admin@example.com",
        )
        self.client.force_login(staff)

        income_response = self.client.get(reverse("admin:ledger_incomecategory_add"))
        self.assertEqual(income_response.status_code, 200)
        self.assertContains(income_response, 'name="category_level"')
        self.assertContains(income_response, 'name="primary_category"')
        self.assertNotContains(income_response, 'name="secondary_category"')
        self.assertNotContains(income_response, 'value="3">三级分类')
        self.assertNotContains(income_response, 'name="parent"')

        expense_response = self.client.get(reverse("admin:ledger_expensecategory_add"))
        self.assertEqual(expense_response.status_code, 200)
        self.assertContains(expense_response, 'name="category_level"')
        self.assertContains(expense_response, 'name="primary_category"')
        self.assertContains(expense_response, 'name="secondary_category"')
        self.assertContains(expense_response, 'value="3">三级分类')
        self.assertContains(expense_response, 'data-family-id="{}"'.format(self.family.pk))
        self.assertContains(expense_response, 'data-parent-id="{}"'.format(primary.pk))
        self.assertContains(expense_response, "js/admin_category_form.js")
        self.assertNotContains(expense_response, 'name="parent"')

    def test_expense_admin_add_form_defaults_family_and_loads_existing_categories(self):
        primary = ExpenseCategory.objects.create(
            family=self.family,
            name="经常性支出",
        )
        secondary = ExpenseCategory.objects.create(
            family=self.family,
            name="餐饮",
            parent=primary,
        )

        form = ExpenseCategoryAdminForm()

        self.assertEqual(form.initial["family"], self.family.pk)
        self.assertQuerySetEqual(
            form.fields["primary_category"].queryset,
            [primary],
        )
        self.assertQuerySetEqual(
            form.fields["secondary_category"].queryset,
            [secondary],
        )
        income_form = IncomeCategoryAdminForm()
        self.assertEqual(
            list(income_form.fields["category_level"].choices),
            [("1", "一级分类"), ("2", "二级分类")],
        )
        self.assertNotIn("secondary_category", income_form.fields)


class LedgerNavigationAndRedirectTests(TestCase):
    def setUp(self):
        self.family = Family.objects.create(name="Navigation Test Family")
        self.member = FamilyMember.objects.create(
            family=self.family,
            display_name="Navigation Tester",
        )
        self.user = get_user_model().objects.create_user(
            username="navigation-tester",
            password="password",
        )
        self.member.user = self.user
        self.member.save(update_fields=["user", "updated_at"])
        self.client.force_login(self.user)

    def test_ledger_page_has_consistent_context_navigation(self):
        response = self.client.get(reverse("ledger:category_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="page-context-nav"')
        self.assertContains(response, 'class="button page-nav-back"')
        self.assertContains(response, "返回上一级")
        self.assertContains(response, reverse("ledger:overview"))
        self.assertContains(response, reverse("dashboard:home"))

    def test_ledger_overview_does_not_repeat_module_home_link(self):
        response = self.client.get(reverse("ledger:overview"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'href="{reverse("dashboard:home")}"',
        )
        self.assertNotContains(response, 'class="button page-nav-module"')

    def test_expense_pages_follow_annual_monthly_record_hierarchy(self):
        monthly_summary = self.client.get(
            reverse("ledger:cashflow_summary_year", args=[2026])
        )
        month_detail = self.client.get(
            reverse(
                "ledger:expense_month_detail",
                kwargs={"year": 2026, "month": 6},
            )
        )
        year_detail = self.client.get(
            reverse("ledger:expense_year_detail", args=[2026])
        )

        self.assertEqual(
            monthly_summary.context["page_parent_url"],
            reverse("ledger:expense_list"),
        )
        expected_detail_parent = reverse(
            "ledger:cashflow_summary_year",
            args=[2026],
        )
        self.assertEqual(
            month_detail.context["page_parent_url"],
            expected_detail_parent,
        )
        self.assertEqual(
            year_detail.context["page_parent_url"],
            expected_detail_parent,
        )
        self.assertNotContains(month_detail, "data-page-back")

    def test_delete_redirect_does_not_accept_scheme_relative_external_url(self):
        record = ExpenseRecord.objects.create(
            family=self.family,
            member=self.member,
            expense_date=date(2026, 6, 28),
            amount=Decimal("10"),
        )

        response = self.client.post(
            reverse("ledger:expense_delete", args=[record.pk]),
            {"next": "//evil.example/steal"},
        )

        self.assertRedirects(
            response,
            reverse("ledger:expense_list"),
            fetch_redirect_response=False,
        )
